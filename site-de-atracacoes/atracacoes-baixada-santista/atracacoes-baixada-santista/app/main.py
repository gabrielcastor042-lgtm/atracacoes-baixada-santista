from __future__ import annotations

import io
import logging
from contextlib import asynccontextmanager
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, List, Optional

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlmodel import select

from .database import get_session, init_db
from .models import Atracacao, SyncStatus
from .scheduler import start_scheduler
from .scrapers.porto_santos import enriquecer_com_rap, fetch_rap_por_navio
from .scrapers.santos_brasil import merge_gate_data, parse_gate_upload, parse_upload
from .sync import contar_ativos, registrar_status, run_gate_sync, run_sync, sincronizar_terminal

logger = logging.getLogger("main")

# (chave da coluna, cabeçalho na planilha)
_EXPORT_COLUMNS = [
    ("rap", "RAP"),
    ("navio", "Navio"),
    ("viagem", "Viagem"),
    ("terminal", "Terminal"),
    ("berco", "Berço"),
    ("eta", "ETA"),
    ("etb", "ETB"),
    ("etd", "ETD"),
    ("ata", "ATA"),
    ("atb", "ATB"),
    ("atd", "ATD"),
    ("deadline_carga", "Deadline carga"),
    ("previsao_abertura_gate", "Previsão abertura gate"),
    ("abertura_gate", "Abertura gate"),
    ("fonte_raw_id", "Fonte (ID original)"),
    ("atualizado_em", "Atualizado em"),
]

_DATE_COLUMNS = {
    "eta", "etb", "etd", "ata", "atb", "atd",
    "deadline_carga", "previsao_abertura_gate", "abertura_gate", "atualizado_em",
}

# Campos de data que podem ser usados no filtro por período (chave = valor
# aceito no parâmetro `campo_data`). Mapear por whitelist em vez de usar
# getattr(Atracacao, campo_data) direto evita expor qualquer atributo do
# modelo (inclusive não-datas) a um parâmetro vindo da URL.
_CAMPO_DATA_MAP = {
    "eta": Atracacao.eta,
    "etb": Atracacao.etb,
    "etd": Atracacao.etd,
    "ata": Atracacao.ata,
    "atb": Atracacao.atb,
    "atd": Atracacao.atd,
    "deadline_carga": Atracacao.deadline_carga,
    "abertura_gate": Atracacao.abertura_gate,
}


def _valor_exportado(row: Atracacao, col: str) -> Any:
    if col == "abertura_gate":
        # Se não tem a abertura efetiva, usa a previsão como referência.
        return row.abertura_gate or row.previsao_abertura_gate
    return getattr(row, col)


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    # A primeira sincronização roda em segundo plano (disparada pelo
    # scheduler), pra não travar a porta do servidor esperando o scraping
    # terminar — isso causava timeout de deploy no Render quando algum
    # terminal estava lento/instável.
    start_scheduler()
    yield


app = FastAPI(title="Atracações - Porto de Santos", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # restrinja em produção
    allow_methods=["*"],
    allow_headers=["*"],
)


def _filtrar_atracacoes(
    q: Optional[str],
    terminal: Optional[str],
    campo_data: Optional[str] = None,
    data_de: Optional[date] = None,
    data_ate: Optional[date] = None,
):
    stmt = select(Atracacao)
    if terminal:
        stmt = stmt.where(Atracacao.terminal == terminal)
    if q:
        like = f"%{q.upper()}%"
        stmt = stmt.where(
            (Atracacao.navio.like(like))  # type: ignore[union-attr]
            | (Atracacao.viagem.like(like))  # type: ignore[union-attr]
        )
    if campo_data and (data_de or data_ate):
        coluna = _CAMPO_DATA_MAP.get(campo_data)
        if coluna is not None:
            if data_de:
                stmt = stmt.where(coluna >= datetime.combine(data_de, time.min))
            if data_ate:
                # time.max inclui o dia final inteiro (até 23:59:59), não só a meia-noite dele.
                stmt = stmt.where(coluna <= datetime.combine(data_ate, time.max))
    return stmt.order_by(Atracacao.eta.is_(None), Atracacao.eta)


@app.get("/buscar", response_model=List[Atracacao])
def buscar(
    q: Optional[str] = Query(None, description="Nome do navio, viagem ou terminal"),
    terminal: Optional[str] = Query(None),
    campo_data: Optional[str] = Query(None, description="Campo de data a filtrar: eta, etb, etd, ata, atb, atd, deadline_carga, abertura_gate"),
    data_de: Optional[date] = Query(None, description="Data inicial (inclusive) do filtro por período"),
    data_ate: Optional[date] = Query(None, description="Data final (inclusive) do filtro por período"),
):
    with get_session() as session:
        stmt = _filtrar_atracacoes(q, terminal, campo_data, data_de, data_ate)
        return session.exec(stmt).all()


@app.get("/exportar")
def exportar(
    q: Optional[str] = Query(None, description="Nome do navio, viagem ou terminal"),
    terminal: Optional[str] = Query(None),
    campo_data: Optional[str] = Query(None),
    data_de: Optional[date] = Query(None),
    data_ate: Optional[date] = Query(None),
):
    """Exporta em Excel (.xlsx) as atracações que batem com o filtro atual
    (mesmos parâmetros da busca)."""
    with get_session() as session:
        stmt = _filtrar_atracacoes(q, terminal, campo_data, data_de, data_ate)
        rows = session.exec(stmt).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Atracações"

    headers = [label for _, label in _EXPORT_COLUMNS]
    ws.append(headers)

    for row in rows:
        ws.append([_valor_exportado(row, col) for col, _ in _EXPORT_COLUMNS])

    date_format = "DD/MM/YYYY HH:MM"
    for col_index, (col, _) in enumerate(_EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_index)
        ws.column_dimensions[letter].width = 20
        if col in _DATE_COLUMNS:
            for cell in ws[letter][1:]:  # pula o cabeçalho
                cell.number_format = date_format

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=atracacoes.xlsx"},
    )


@app.post("/sync")
def sync_now():
    """Dispara uma sincronização manual (útil para testes/depuração) —
    inclui também a atualização de gate/deadline da Santos Brasil."""
    resultado = run_sync()
    resultado["santos_brasil_gate_atualizados"] = run_gate_sync()
    return resultado


@app.post("/upload/santos_brasil")
async def upload_santos_brasil(
    arquivo_excel: UploadFile = File(..., description="Lista de Atracação (.xls)"),
    arquivo_gate: Optional[UploadFile] = File(
        None,
        description=(
            "Lista de Recebimento salva como HTML (opcional — o gate já é "
            "atualizado sozinho a cada 30 min via API; envie só se quiser "
            "reforçar/sobrescrever na hora)."
        ),
    ),
):
    """Recebe a planilha (.xls) principal exportada manualmente da área de
    cliente da Santos Brasil e sincroniza os navios no banco. O relatório
    de liberação de gate (arquivo_gate) é opcional — o gate/deadline já é
    mantido atualizado em segundo plano via `fetch_gate_data` (ver
    scrapers/santos_brasil.py); só processamos esse arquivo se ele vier
    junto."""
    try:
        conteudo_excel = await arquivo_excel.read()
        records = parse_upload(conteudo_excel)
        if not records:
            raise HTTPException(422, "Nenhum navio encontrado na planilha (.xls) enviada.")

        if arquivo_gate is not None:
            conteudo_gate = await arquivo_gate.read()
            gate_por_id = parse_gate_upload(conteudo_gate)
            records = merge_gate_data(records, gate_por_id)

        try:
            rap_lookup = fetch_rap_por_navio()
            enriquecer_com_rap(records, "santos_brasil", rap_lookup)
        except Exception:
            pass  # RAP é um complemento — não impede o upload de seguir

        with get_session() as session:
            aviso = sincronizar_terminal(session, "santos_brasil", records)
            session.commit()
            total_atual = contar_ativos(session, "santos_brasil")
            if aviso:
                registrar_status(session, "santos_brasil", erro=aviso)
            else:
                registrar_status(session, "santos_brasil", total_atual)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        # Sem isso, um erro inesperado aqui vira uma resposta HTML/texto
        # puro do próprio servidor ("Internal Server Error"), que o
        # navegador não consegue interpretar como JSON — a tela só
        # mostrava um erro genérico e ilegível de parsing.
        logger.exception("Falha ao processar upload da Santos Brasil")
        raise HTTPException(500, f"Erro ao processar os arquivos: {exc}") from exc

    com_gate = sum(1 for r in records if r.get("abertura_gate") or r.get("previsao_abertura_gate"))
    return {"terminal": "santos_brasil", "registros": total_atual, "com_gate": com_gate}


@app.get("/status", response_model=List[SyncStatus])
def status():
    """Quando cada fonte (scraper automático ou upload manual) foi
    sincronizada pela última vez."""
    with get_session() as session:
        return session.exec(select(SyncStatus)).all()


@app.get("/health")
def health():
    return {"status": "ok"}


app.mount("/", StaticFiles(directory=Path(__file__).parent / "static", html=True), name="static")